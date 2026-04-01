import logging
from datetime import date
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(bold=True, color="276221")
NO_FILL = PatternFill(fill_type=None)
NORMAL_FONT = Font()


class ExcelService:
    def __init__(self):
        self.output_path = Path(settings.EXCEL_OUTPUT_PATH)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def _get_or_create_workbook(self):
        if self.output_path.exists():
            return load_workbook(self.output_path)
        wb = Workbook()
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        return wb

    def _apply_header_style(self, cell):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _apply_border(self, cell):
        thin = Side(style="thin")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _get_or_create_monthly_sheet(self, wb, sheet_name: str):
        """Obtiene o crea la hoja del mes actual. Se crea vacía sin formato."""
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        return ws

    def _mark_best_prices(self, ws, first_provider_col: int):
        """
        Marca en verde el precio más bajo de cada fila de datos.
        Itera TODAS las filas (incluyendo múltiples días) — cada fila se evalúa
        independientemente para que cada día tenga su propio marcado correcto.
        """
        for row in range(3, ws.max_row + 1):
            if not ws.cell(row, 2).value:  # Col B = Medicamento
                continue

            prices = []
            col = first_provider_col
            while col <= ws.max_column:
                val = ws.cell(row, col).value
                if isinstance(val, (int, float)):
                    prices.append((col, val))
                col += 2  # Saltar columna de cantidad

            if len(prices) < 2:
                continue

            min_price = min(v for _, v in prices)

            col = first_provider_col
            while col <= ws.max_column:
                cell = ws.cell(row, col)
                if isinstance(cell.value, (int, float)) and cell.value == min_price:
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT
                else:
                    cell.fill = NO_FILL
                    cell.font = NORMAL_FONT
                col += 2

    def update_prices(self, provider_name: str, prices: list[dict], report_date: date = None) -> str:
        """
        Actualiza el archivo Excel con los precios del proveedor.
        Formato: Fila 1 = proveedor (fusionado Precio+Cantidad), Fila 2 = encabezados,
                 Fila 3+ = datos. Col A=Fecha, B=Medicamento, C/D=Prov1, E/F=Prov2, ...
        """
        if not prices:
            logger.info(f"No hay precios para actualizar de proveedor: {provider_name}")
            return str(self.output_path)

        report_date = report_date or date.today()
        sheet_name = report_date.strftime("%Y-%m")

        wb = self._get_or_create_workbook()

        # Obtener o crear hoja en el nuevo formato
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_name)
            ws.cell(2, 1, "Fecha")
            self._apply_header_style(ws.cell(2, 1))
            ws.cell(2, 2, "Medicamento")
            self._apply_header_style(ws.cell(2, 2))
            ws.column_dimensions["A"].width = 13
            ws.column_dimensions["B"].width = 36
        else:
            ws = wb[sheet_name]

        # Buscar o crear las columnas del proveedor
        # Fila 1: nombre proveedor en cols impares (3, 5, 7...), paso de 2 en 2
        first_provider_col = 3
        provider_price_col = None
        col = first_provider_col
        while True:
            cell_val = ws.cell(1, col).value
            if cell_val == provider_name:
                provider_price_col = col
                break
            if cell_val is None:
                # Nueva columna de proveedor
                provider_price_col = col
                unit_col = col + 1
                prov_cell = ws.cell(1, col, provider_name)
                self._apply_header_style(prov_cell)
                ws.merge_cells(
                    start_row=1, start_column=col,
                    end_row=1, end_column=unit_col,
                )
                ws.cell(2, col, "Precio")
                self._apply_header_style(ws.cell(2, col))
                ws.cell(2, unit_col, "Cantidad")
                self._apply_header_style(ws.cell(2, unit_col))
                ws.column_dimensions[get_column_letter(col)].width = 16
                ws.column_dimensions[get_column_letter(unit_col)].width = 14
                break
            col += 2

        provider_unit_col = provider_price_col + 1

        date_str = report_date.strftime("%d/%m/%Y")

        # Todos los medicamentos (cualquier fecha) — para validación y posición final
        all_med_rows = {}
        for row in range(3, ws.max_row + 1):
            val = ws.cell(row, 2).value
            if val and isinstance(val, str):
                all_med_rows[val.lower()] = row

        # Solo medicamentos de HOY — únicamente estas filas se actualizan
        today_med_rows = {}
        for row in range(3, ws.max_row + 1):
            fecha_val = ws.cell(row, 1).value
            med_val = ws.cell(row, 2).value
            if med_val and isinstance(med_val, str) and fecha_val == date_str:
                today_med_rows[med_val.lower()] = row

        def find_med_key(search_key: str, source: dict) -> str | None:
            from difflib import SequenceMatcher
            if search_key in source:
                return search_key
            search_words = set(search_key.split())
            best_word, best_overlap = None, 0
            for k in source:
                kw = set(k.split())
                if search_words.issubset(kw) or kw.issubset(search_words):
                    overlap = len(search_words & kw)
                    if overlap > best_overlap:
                        best_overlap, best_word = overlap, k
            if best_word:
                return best_word
            best_fuzzy, best_ratio = None, 0.0
            for k in source:
                r = SequenceMatcher(None, search_key, k).ratio()
                if r >= 0.80 and r > best_ratio:
                    best_ratio, best_fuzzy = r, k
            return best_fuzzy

        # Siguiente fila disponible = después de todos los datos existentes
        next_available_row = (max(all_med_rows.values()) + 1) if all_med_rows else 3

        for item in prices:
            med_name = item.get("medication_name", "").strip()
            price_val = item.get("price")
            unit = item.get("unit") or "NO"

            if not med_name or price_val is None:
                continue

            key = med_name.lower()

            # 1. Buscar en las filas de HOY
            matched_key = find_med_key(key, today_med_rows)

            if matched_key is None:
                # Medicamento no está en las filas de hoy → ignorar completamente.
                # Solo se actualiza lo que el admin ya inicializó para este día.
                logger.info(f"Excel: '{med_name}' no está en las filas de hoy ({date_str}), se ignora")
                continue

            row_idx = today_med_rows[matched_key]
            # Actualizar fecha
            ws.cell(row_idx, 1, date_str)
            self._apply_border(ws.cell(row_idx, 1))
            # Precio del proveedor
            price_cell = ws.cell(row_idx, provider_price_col, price_val)
            price_cell.number_format = "#,##0.00"
            self._apply_border(price_cell)
            # Cantidad del proveedor
            unit_cell = ws.cell(row_idx, provider_unit_col, unit)
            self._apply_border(unit_cell)

        # Marcar mejor precio en verde
        self._mark_best_prices(ws, first_provider_col)

        wb.save(self.output_path)
        logger.info(
            f"Excel actualizado: hoja '{sheet_name}', proveedor '{provider_name}', {len(prices)} precios"
        )
        return str(self.output_path)

    def generate_report(self, prices_data: list[dict]) -> bytes:
        """
        Genera Excel desde datos de la BD con el mismo formato que Google Sheets.
        prices_data: lista de {medication_name, price, unit, provider_name, date_reported}
        Una hoja por mes. Formato: Fecha | Medicamento | Precio+Cantidad por proveedor.
        """
        from io import BytesIO

        wb = Workbook()
        wb.remove(wb.active)

        # Agrupar por mes
        months = sorted(set(p["date_reported"].strftime("%Y-%m") for p in prices_data))

        for month in months:
            month_prices = [p for p in prices_data if p["date_reported"].strftime("%Y-%m") == month]
            ws = wb.create_sheet(title=month)

            # Orden de proveedores según primera aparición
            providers = []
            for p in month_prices:
                if p["provider_name"] not in providers:
                    providers.append(p["provider_name"])

            # Col A=Fecha, Col B=Medicamento, luego 2 cols por proveedor
            ws.cell(2, 1, "Fecha"); self._apply_header_style(ws.cell(2, 1))
            ws.cell(2, 2, "Medicamento"); self._apply_header_style(ws.cell(2, 2))
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 38

            provider_cols = {}
            for i, prov in enumerate(providers):
                price_col = 3 + i * 2
                unit_col = price_col + 1
                provider_cols[prov] = price_col

                # Fila 2: Precio, Cantidad
                ws.cell(2, price_col, "Precio"); self._apply_header_style(ws.cell(2, price_col))
                ws.cell(2, unit_col, "Cantidad"); self._apply_header_style(ws.cell(2, unit_col))
                ws.column_dimensions[get_column_letter(price_col)].width = 16
                ws.column_dimensions[get_column_letter(unit_col)].width = 14

                # Fila 1: nombre del proveedor combinado
                if price_col == 3:
                    ws.cell(1, 1, prov)
                    self._apply_header_style(ws.cell(1, 1))
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=unit_col)
                else:
                    ws.cell(1, price_col, prov)
                    self._apply_header_style(ws.cell(1, price_col))
                    ws.merge_cells(start_row=1, start_column=price_col, end_row=1, end_column=unit_col)

            # Pivot: {med_key: {name, date, provider_name: {price, unit}}}
            # Ordenar por fecha para que el último precio prevalezca en duplicados
            pivot = {}
            med_order = []
            for p in sorted(month_prices, key=lambda x: x["date_reported"]):
                key = p["medication_name"].lower()
                if key not in pivot:
                    pivot[key] = {"name": p["medication_name"], "date": p["date_reported"]}
                    med_order.append(key)
                pivot[key][p["provider_name"]] = {
                    "price": p["price"],
                    "unit": p["unit"] or "NO",
                    "date": p["date_reported"],
                }

            # Escribir filas de datos
            for row_num, key in enumerate(med_order, start=3):
                data = pivot[key]

                # Fecha: la más reciente entre proveedores
                latest_date = max(
                    (data[pn]["date"] for pn in providers if pn in data),
                    default=data["date"],
                )
                date_cell = ws.cell(row_num, 1, latest_date.strftime("%d/%m/%Y"))
                self._apply_border(date_cell)

                med_cell = ws.cell(row_num, 2, data["name"])
                self._apply_border(med_cell)

                provider_prices = []
                for prov_name in providers:
                    price_col = provider_cols[prov_name]
                    unit_col = price_col + 1
                    if prov_name in data:
                        price_cell = ws.cell(row_num, price_col, data[prov_name]["price"])
                        price_cell.number_format = "#,##0.00"
                        self._apply_border(price_cell)
                        provider_prices.append((price_col, data[prov_name]["price"]))
                        unit_cell = ws.cell(row_num, unit_col, data[prov_name]["unit"])
                        self._apply_border(unit_cell)
                    else:
                        ws.cell(row_num, price_col, "NO"); self._apply_border(ws.cell(row_num, price_col))
                        ws.cell(row_num, unit_col, "NO"); self._apply_border(ws.cell(row_num, unit_col))

                # Marcar precio más bajo en verde
                if len(provider_prices) >= 2:
                    min_price = min(v for _, v in provider_prices)
                    for col, val in provider_prices:
                        cell = ws.cell(row_num, col)
                        if val == min_price:
                            cell.fill = GREEN_FILL
                            cell.font = GREEN_FONT
                        else:
                            cell.fill = NO_FILL
                            cell.font = NORMAL_FONT

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def create_empty_monthly_sheet(
        self,
        provider_names: list[str],
        target_date: date = None,
        force: bool = False,
    ) -> bool:
        """
        Crea la hoja del mes como pestaña vacía (sin formato ni encabezados).
        El formato se aplica cuando el usuario pulsa "Inicializar día".
        Si force=True elimina la hoja existente y la recrea vacía.
        Retorna True si se creó, False si ya existía (y no se forzó).
        """
        target_date = target_date or date.today()
        sheet_name = target_date.strftime("%Y-%m")

        wb = self._get_or_create_workbook()

        if sheet_name in wb.sheetnames:
            if not force:
                logger.info(f"Hoja '{sheet_name}' ya existe, no se sobreescribe")
                return False
            del wb[sheet_name]
            logger.info(f"Hoja '{sheet_name}' eliminada para recrear")

        wb.create_sheet(title=sheet_name)
        wb.save(self.output_path)
        logger.info(f"Hoja '{sheet_name}' creada (vacía, sin formato)")
        return True

    def create_day_header(
        self, target_date: date = None, provider_names: list[str] | None = None
    ) -> dict:
        """
        Inserta el bloque de encabezado del nuevo día en el Excel local.
        Si la hoja está vacía (primer día del mes), crea la estructura completa
        (filas 1-2 con proveedores y sub-encabezados) usando provider_names.
        Para días posteriores añade un nuevo bloque debajo de los datos existentes.
        Retorna {"created": bool}.
        """
        target_date = target_date or date.today()
        sheet_name = target_date.strftime("%Y-%m")

        wb = self._get_or_create_workbook()
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)

        ws = wb[sheet_name]
        date_str = target_date.strftime("%d/%m/%Y")
        MARKER_COL = 26  # Columna Z — invisible para el usuario

        # Buscar marcador desde fila 2 para cubrir también el primer día
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, MARKER_COL).value
            if val == date_str:
                logger.info(f"Excel: encabezado del día '{date_str}' ya existe en fila {row}, no se duplica")
                return {"created": False}

        # Detectar si la hoja está vacía (sin proveedores en fila 1)
        price_cols = []
        col = 3
        while col <= ws.max_column:
            if ws.cell(1, col).value:
                price_cols.append(col)
            col += 2
        is_blank = not price_cols

        if is_blank:
            # Primer día del mes: crear estructura inicial (filas 1-2) con los proveedores
            if not provider_names:
                logger.warning("Excel: hoja vacía pero no se proporcionaron proveedores")
                return {"created": False}

            ws.cell(2, 1, "Fecha")
            self._apply_header_style(ws.cell(2, 1))
            ws.cell(2, 2, "Medicamento")
            self._apply_header_style(ws.cell(2, 2))
            ws.column_dimensions["A"].width = 13
            ws.column_dimensions["B"].width = 36

            for i, prov_name in enumerate(provider_names):
                price_col = 3 + i * 2
                unit_col = price_col + 1
                prov_cell = ws.cell(1, price_col, prov_name)
                self._apply_header_style(prov_cell)
                ws.merge_cells(start_row=1, start_column=price_col, end_row=1, end_column=unit_col)
                ws.cell(2, price_col, "Precio")
                self._apply_header_style(ws.cell(2, price_col))
                ws.cell(2, unit_col, "Cantidad")
                self._apply_header_style(ws.cell(2, unit_col))
                ws.column_dimensions[get_column_letter(price_col)].width = 16
                ws.column_dimensions[get_column_letter(unit_col)].width = 14

            # Marcador en col Z de fila 2
            ws.cell(2, MARKER_COL, date_str)
            wb.save(self.output_path)
            logger.info(
                f"Excel: estructura inicial del mes creada con {len(provider_names)} proveedor(es)"
            )
            return {"created": True}

        # Día posterior: añadir nuevo bloque debajo de los datos existentes
        last_data_row = 2
        for row in range(3, ws.max_row + 1):
            if ws.cell(row, 1).value or ws.cell(row, 2).value:
                last_data_row = row

        prov_header_row = last_data_row + 2
        col_header_row = prov_header_row + 1

        for pc in price_cols:
            uc = pc + 1
            prov_name = ws.cell(1, pc).value or ""
            if pc == 3:
                prov_cell = ws.cell(prov_header_row, 1, prov_name)
                self._apply_header_style(prov_cell)
                try:
                    ws.merge_cells(
                        start_row=prov_header_row, start_column=1,
                        end_row=prov_header_row, end_column=uc,
                    )
                except Exception:
                    pass
            else:
                prov_cell = ws.cell(prov_header_row, pc, prov_name)
                self._apply_header_style(prov_cell)
                try:
                    ws.merge_cells(
                        start_row=prov_header_row, start_column=pc,
                        end_row=prov_header_row, end_column=uc,
                    )
                except Exception:
                    pass

        ws.cell(col_header_row, MARKER_COL, date_str)
        ws.cell(col_header_row, 1, "Fecha")
        self._apply_header_style(ws.cell(col_header_row, 1))
        ws.cell(col_header_row, 2, "Medicamento")
        self._apply_header_style(ws.cell(col_header_row, 2))
        for pc in price_cols:
            ws.cell(col_header_row, pc, "Precio")
            self._apply_header_style(ws.cell(col_header_row, pc))
            ws.cell(col_header_row, pc + 1, "Cantidad")
            self._apply_header_style(ws.cell(col_header_row, pc + 1))

        wb.save(self.output_path)
        logger.info(f"Excel: encabezado de nuevo día insertado en filas {prov_header_row}-{col_header_row}")
        return {"created": True}

    def monthly_sheet_exists(self, target_date: date = None) -> bool:
        """Verifica si ya existe la hoja del mes indicado."""
        target_date = target_date or date.today()
        sheet_name = target_date.strftime("%Y-%m")
        if not self.output_path.exists():
            return False
        from openpyxl import load_workbook
        wb = load_workbook(self.output_path, read_only=True)
        exists = sheet_name in wb.sheetnames
        wb.close()
        return exists

    def get_summary(self) -> dict:
        if not self.output_path.exists():
            return {"exists": False, "medications": 0, "providers": []}

        wb = load_workbook(self.output_path)
        current_sheet = date.today().strftime("%Y-%m")

        if current_sheet not in wb.sheetnames:
            return {"exists": True, "medications": 0, "providers": [], "sheets": wb.sheetnames}

        ws = wb[current_sheet]
        providers = []
        for col in range(2, ws.max_column + 1):
            val = ws.cell(1, col).value
            if val:
                providers.append(val)

        medications = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row, 2).value)

        return {
            "exists": True,
            "current_sheet": current_sheet,
            "medications": medications,
            "providers": providers,
            "file_path": str(self.output_path),
            "all_sheets": wb.sheetnames,
        }
